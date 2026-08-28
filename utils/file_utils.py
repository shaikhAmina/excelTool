import streamlit as st
import pandas as pd
import zipfile
from io import BytesIO
from openpyxl.styles import PatternFill

from utils.helpers import (
    show_file_info,
    show_file_info_multi,
    show_data_preview,
    result_banner,
    friendly_error,
    df_to_excel_bytes,
    privacy_notice,
    feedback_widget,
)


# ─────────────────────────────────────────────────────
# MERGE EXCEL
# ─────────────────────────────────────────────────────
def merge_excel_ui():
    st.markdown("### 🧩 Merge Excel Files")
    st.caption("Combine multiple Excel files — or all sheets inside one file — into a single spreadsheet.")
    privacy_notice()
    st.markdown("---")

    merge_type = st.radio(
        "What would you like to merge?",
        ["📁 Multiple Excel Files → one sheet", "📑 All Sheets in one file → one sheet"],
        horizontal=True,
    )

    # ── STEP 1: Upload ────────────────────────────────
    st.markdown("#### Step 1 — Upload")

    if merge_type == "📁 Multiple Excel Files → one sheet":
        files = st.file_uploader(
            "Upload Excel files (select all at once)",
            type=["xlsx"],
            accept_multiple_files=True,
            key="merge_multi",
        )
        if not files:
            st.info("Upload two or more .xlsx files to get started.")
            feedback_widget()
            return

        show_file_info_multi(files)

        # ── STEP 2: Configure ─────────────────────────
        st.markdown("#### Step 2 — Configure")
        add_source_col = st.checkbox(
            "Add a 'Source File' column so you can trace each row back to its original file",
            value=True,
        )

        # ── STEP 3: Preview first file ────────────────
        st.markdown("#### Step 3 — Preview")
        try:
            preview_df = pd.read_excel(files[0])
            show_data_preview(preview_df, label=f"Preview of '{files[0].name}'")
        except Exception as e:
            friendly_error(e)
            return

        # ── STEP 4: Process ───────────────────────────
        st.markdown("#### Step 4 — Process")
        if st.button("🧩 Merge Files", type="primary"):
            try:
                frames = []
                for f in files:
                    df = pd.read_excel(f)
                    if add_source_col:
                        df["Source_File"] = f.name
                    frames.append(df)

                merged = pd.concat(frames, ignore_index=True)
                total_rows = len(merged)

                # ── STEP 5: Result ────────────────────
                result_banner(
                    "Merge completed",
                    {
                        "Files processed": len(files),
                        "Total rows": f"{total_rows:,}",
                        "Columns": len(merged.columns),
                    },
                )
                show_data_preview(merged, label="Merged result")

                # ── STEP 6: Download ──────────────────
                st.download_button(
                    "⬇️ Download Merged Excel",
                    df_to_excel_bytes(merged, "Merged_Data"),
                    file_name="merged_files.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                friendly_error(e)

    else:  # Merge sheets
        file = st.file_uploader(
            "Upload an Excel file with multiple sheets",
            type=["xlsx"],
            key="merge_sheets",
        )
        if not file:
            st.info("Upload an .xlsx file that contains multiple sheets.")
            feedback_widget()
            return

        # ── STEP 2: Sheet info ────────────────────────
        try:
            sheet_dict = pd.read_excel(file, sheet_name=None)
        except Exception as e:
            friendly_error(e)
            return

        sheet_names = list(sheet_dict.keys())
        st.markdown("#### Step 2 — Configure")
        st.write(f"Found **{len(sheet_names)} sheets**: {', '.join(sheet_names)}")
        selected_sheets = st.multiselect(
            "Choose which sheets to merge (leave blank = all)",
            sheet_names,
            default=sheet_names,
        )
        add_sheet_col = st.checkbox("Add a 'Source_Sheet' column", value=True)

        # ── STEP 3: Preview ───────────────────────────
        st.markdown("#### Step 3 — Preview")
        first_sheet = selected_sheets[0] if selected_sheets else sheet_names[0]
        show_data_preview(sheet_dict[first_sheet], label=f"Sheet '{first_sheet}'")

        # ── STEP 4: Process ───────────────────────────
        st.markdown("#### Step 4 — Process")
        if st.button("🧩 Merge Sheets", type="primary"):
            try:
                frames = []
                for s in (selected_sheets or sheet_names):
                    df = sheet_dict[s].copy()
                    if add_sheet_col:
                        df["Source_Sheet"] = s
                    frames.append(df)

                merged = pd.concat(frames, ignore_index=True)

                result_banner(
                    "Merge completed",
                    {
                        "Sheets merged": len(frames),
                        "Total rows": f"{len(merged):,}",
                        "Columns": len(merged.columns),
                    },
                )
                show_data_preview(merged, label="Merged result")

                st.download_button(
                    "⬇️ Download Merged Excel",
                    df_to_excel_bytes(merged, "Merged_Sheets"),
                    file_name="merged_sheets.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                friendly_error(e)

    feedback_widget()


# ─────────────────────────────────────────────────────
# SPLIT EXCEL
# ─────────────────────────────────────────────────────
def split_excel_ui():
    st.markdown("### ✂️ Split Excel File")
    st.caption("Break a large spreadsheet into smaller files and download them as a ZIP.")
    privacy_notice()
    st.markdown("---")

    # ── STEP 1: Upload ────────────────────────────────
    st.markdown("#### Step 1 — Upload")
    file = st.file_uploader("Upload an Excel file (.xlsx)", type=["xlsx"], key="split_upload")
    if not file:
        st.info("Upload an .xlsx file to get started.")
        feedback_widget()
        return

    try:
        df = pd.read_excel(file)
    except Exception as e:
        friendly_error(e)
        return

    show_file_info(file, df)

    # ── STEP 2: Configure ─────────────────────────────
    st.markdown("#### Step 2 — Configure")
    entries_per_file = st.number_input(
        "Rows per output file",
        min_value=1,
        max_value=len(df),
        value=min(500, len(df)),
        step=100,
        help="Each output file will contain at most this many data rows.",
    )
    num_parts = -(-len(df) // entries_per_file)  # ceiling division
    st.caption(f"This will produce **{num_parts} file(s)**.")

    # ── STEP 3: Preview ───────────────────────────────
    st.markdown("#### Step 3 — Preview")
    show_data_preview(df, label="Source file")

    # ── STEP 4: Process ───────────────────────────────
    st.markdown("#### Step 4 — Process")
    if st.button("✂️ Split File", type="primary"):
        try:
            chunks = [df.iloc[i: i + entries_per_file] for i in range(0, len(df), entries_per_file)]
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for idx, chunk in enumerate(chunks, 1):
                    zf.writestr(f"part_{idx}.xlsx", df_to_excel_bytes(chunk.reset_index(drop=True)))

            result_banner(
                "Split completed",
                {
                    "Output files": len(chunks),
                    "Rows per file": f"up to {entries_per_file:,}",
                    "Source rows": f"{len(df):,}",
                },
            )

            st.download_button(
                "⬇️ Download All Parts (ZIP)",
                zip_buffer.getvalue(),
                file_name=f"split_{file.name.replace('.xlsx','')}_{len(chunks)}_parts.zip",
                mime="application/zip",
            )
        except Exception as e:
            friendly_error(e)

    feedback_widget()


# ─────────────────────────────────────────────────────
# FIND DUPLICATES
# ─────────────────────────────────────────────────────
def find_duplicates_ui():
    st.markdown("### 🔍 Find Duplicates")
    st.caption("Detect duplicate rows and download a highlighted Excel file.")
    privacy_notice()
    st.markdown("---")

    # ── STEP 1: Upload ────────────────────────────────
    st.markdown("#### Step 1 — Upload")
    file = st.file_uploader("Upload an Excel file (.xlsx)", type=["xlsx"], key="dup_upload")
    if not file:
        st.info("Upload an .xlsx file to get started.")
        feedback_widget()
        return

    try:
        df = pd.read_excel(file)
    except Exception as e:
        friendly_error(e)
        return

    show_file_info(file, df)

    # ── STEP 2: Configure ─────────────────────────────
    st.markdown("#### Step 2 — Configure")
    cols = st.multiselect(
        "Which columns should be checked for duplicates?",
        list(df.columns),
        default=list(df.columns[:1]) if len(df.columns) else [],
        help="Rows are considered duplicates when ALL selected columns match.",
    )
    keep_option = st.radio(
        "Which duplicates to mark?",
        ["All occurrences", "Keep first — mark rest", "Keep last — mark rest"],
        horizontal=True,
    )
    keep_map = {
        "All occurrences": False,
        "Keep first — mark rest": "first",
        "Keep last — mark rest": "last",
    }

    # ── STEP 3: Preview ───────────────────────────────
    st.markdown("#### Step 3 — Preview")
    show_data_preview(df, label="Source file")

    # ── STEP 4: Process ───────────────────────────────
    st.markdown("#### Step 4 — Process")
    if not cols:
        st.warning("Select at least one column above before running.")
        feedback_widget()
        return

    if st.button("🔍 Find Duplicates", type="primary"):
        try:
            keep_val = keep_map[keep_option]
            dups_mask = df.duplicated(subset=cols, keep=keep_val)
            dup_df = df[dups_mask]
            dup_count = int(dups_mask.sum())

            result_banner(
                "Scan complete",
                {
                    "Duplicate rows found": f"{dup_count:,}",
                    "Unique rows": f"{len(df) - dup_count:,}",
                    "Columns checked": ", ".join(cols),
                },
            )

            if dup_count > 0:
                show_data_preview(dup_df.reset_index(drop=True), label="Duplicate rows")

                # Build highlighted Excel
                output = BytesIO()
                yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Sheet1")
                    sheet = writer.sheets["Sheet1"]
                    for col_name in cols:
                        col_idx = df.columns.get_loc(col_name) + 1
                        dup_vals = set(
                            df[df.duplicated(subset=[col_name], keep=keep_val)][col_name].tolist()
                        )
                        for r_idx, val in enumerate(df[col_name], start=2):
                            if val in dup_vals:
                                sheet.cell(row=r_idx, column=col_idx).fill = yellow

                st.download_button(
                    "⬇️ Download Highlighted Excel",
                    output.getvalue(),
                    file_name=f"duplicates_{file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.success("🎉 No duplicates found in the selected columns!")

        except Exception as e:
            friendly_error(e)

    feedback_widget()
